"""
Genera prestadores_servy.xlsx listo para cargar datos reales.
Uso: python generar_excel_prestadores.py
"""
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

# ── 120+ categorias con descripcion ──────────────────────────────────────────
CATEGORIAS = [
  # PLOMERIA
  ("Plomero",                       "Reparacion de canarias, perdidas, canillas, termotanques y calefones"),
  ("Plomero sanitario",             "Instalacion y reparacion de banos, inodoros, piletas y duchas"),
  ("Destapaciones",                 "Desobstruccion de canarias, cloacas, desagues y piletas tapadas"),
  ("Instalador de termotanque",     "Colocacion y conexion de termotanques y calefones a gas/electrico"),
  ("Plomero de piletas",            "Instalacion y reparacion de sistemas hidraulicos de piletas de natacion"),
  ("Plomero de riego",              "Diseno e instalacion de sistemas de riego automatico para jardines"),
  ("Impermeabilizacion",            "Aplicacion de membranas y productos impermeabilizantes en techos y paredes"),
  # ELECTRICIDAD
  ("Electricista",                  "Instalaciones electricas domiciliarias, reparaciones y tableros"),
  ("Electricista industrial",       "Instalaciones electricas trifasicas, motores y tableros industriales"),
  ("Electricista de exteriores",    "Iluminacion exterior, jardines, cocheras y frentes de edificios"),
  ("Instalador de luces LED",       "Colocacion de tiras LED, spots y sistemas de iluminacion eficiente"),
  ("Instalador de tableros electricos", "Armado y actualizacion de tableros domiciliarios e industriales"),
  ("Instalador de paneles solares", "Montaje de sistemas fotovoltaicos para generacion de energia solar"),
  ("Tecnico en domótica",           "Automatizacion del hogar: luces, persianas, temperatura y seguridad"),
  ("Instalador de UPS",             "Colocacion de sistemas de energia ininterrumpida para equipos criticos"),
  # GAS
  ("Gasista matriculado",           "Instalacion y reparacion de redes de gas con habilitacion oficial"),
  ("Tecnico en calefones",          "Reparacion y mantenimiento de calefones a gas y electricos"),
  ("Tecnico en hornallas",          "Reparacion de cocinas a gas, hornallas y hornos"),
  ("Instalador de estufas",         "Colocacion y conexion de estufas a gas, pellet y lena"),
  ("Instalador de calefaccion central", "Instalacion de sistemas de calefaccion central por agua caliente"),
  ("Tecnico en generadores",        "Mantenimiento y reparacion de grupos electrogenos a gas y nafta"),
  # ALBANILERIA Y CONSTRUCCION
  ("Albanil",                       "Construccion, refacciones, mamposteria y trabajos de obra general"),
  ("Constructor",                   "Direccion y ejecucion de obras de construccion y ampliacion"),
  ("Yesero",                        "Aplicacion de yeso en paredes, cielorrasos y remodelaciones interiores"),
  ("Instalador de Durlock",         "Colocacion de placas de roca de yeso, tabiques y cielorrasos en seco"),
  ("Instalador de cielorraso",      "Colocacion de cielorrasos de yeso, PVC, madera y desmontables"),
  ("Techista",                      "Reparacion e instalacion de techos de chapa, tejas y membrana"),
  ("Colocador de tejas",            "Instalacion de techos con tejas ceramicas, francesas o coloniales"),
  ("Colocador de chapas",           "Colocacion de techos de chapa galvanizada, sinusoidal y trapezoidal"),
  ("Instalador de pisos",           "Colocacion de pisos de ceramica, porcelanato, piso flotante y madera"),
  ("Instalador de ceramicos",       "Colocacion de ceramicos y azulejos en paredes y pisos"),
  ("Instalador de porcelanato",     "Colocacion de piso y revestimiento de porcelanato"),
  ("Instalador de piso flotante",   "Colocacion de piso laminado y flotante de madera o HDF"),
  ("Instalador de madera maciza",   "Colocacion de piso de madera maciza, parquet y deck interior"),
  ("Instalador de pisos vinilicos", "Colocacion de pisos vinilicos en rollo o tablones"),
  ("Instalador de pisos de goma",   "Colocacion de pisos de goma para gimnasios, garajes y patios"),
  ("Instalador de pisos industriales", "Aplicacion de pisos de cemento alisado, epoxi y poliuretano"),
  ("Colocador de mosaicos",         "Colocacion de mosaicos, microcemento y venecitas"),
  # PINTURA
  ("Pintor",                        "Pintura interior y exterior de viviendas y edificios"),
  ("Pintor de interiores",          "Pintura de ambientes interiores, techos y molduras"),
  ("Pintor de exteriores",          "Pintura de frentes, balcones, medianeras y exterior"),
  ("Pintor de frentes",             "Renovacion y pintura de fachadas de viviendas y comercios"),
  # CARPINTERIA Y MADERA
  ("Carpintero",                    "Fabricacion y reparacion de muebles, puertas y aberturas de madera"),
  ("Carpintero de muebles",         "Fabricacion a medida de placares, estantes, mesas y sillas"),
  ("Carpintero de puertas",         "Instalacion y reparacion de puertas de madera y marcos"),
  ("Carpintero de exteriores",      "Deck de madera, pergolas, cercos y estructuras exteriores en madera"),
  ("Instalador de deck",            "Colocacion de pisos de deck en madera o composite para exteriores"),
  ("Instalador de placares",        "Fabricacion e instalacion de placares y armarios empotrados"),
  ("Instalador de muebles de cocina", "Colocacion y montaje de muebles de cocina y mesadas"),
  ("Instalador de pergolas",        "Construccion de pergolas de madera, aluminio y hierro para jardines"),
  # HERRERIA Y METALURGIA
  ("Herrero",                       "Fabricacion e instalacion de rejas, portones y estructuras metalicas"),
  ("Soldador",                      "Soldadura MIG, TIG y electrodo para estructuras y reparaciones metalicas"),
  ("Soldador TIG/MIG",              "Soldadura de precision para acero inoxidable, aluminio y acero al carbono"),
  ("Tornero",                       "Mecanizado de piezas metalicas en torno convencional o CNC"),
  ("Herreria artistica",            "Fabricacion de rejas ornamentales, barandas y herreria decorativa"),
  ("Herreria industrial",           "Estructuras metalicas para industria, galpones y construccion"),
  ("Instalador de portones electricos", "Colocacion de portones seccionales y corredizos con motor"),
  ("Instalador de rejas",           "Instalacion de rejas de seguridad en ventanas, puertas y garajes"),
  ("Instalador de cercos",          "Colocacion de cercos metalicos, alambrados y mallas de seguridad"),
  # CERRAJERIA
  ("Cerrajero",                     "Apertura de puertas, cambio de cerraduras y duplicado de llaves"),
  ("Instalador de cerradura digital", "Colocacion de cerraduras electronicas y biometricas"),
  ("Instalador de cajas fuertes",   "Instalacion y apertura de cajas de seguridad embutidas y de pie"),
  # LIMPIEZA
  ("Limpieza del hogar",            "Limpieza general de viviendas, departamentos y oficinas"),
  ("Mucama",                        "Servicio de limpieza y orden domestico periodico"),
  ("Limpieza post obra",            "Limpieza profunda de inmuebles luego de refacciones o construccion"),
  ("Limpieza de alfombras",         "Lavado y sanitizacion de alfombras, tapetes y moquettes"),
  ("Limpieza de tapizados",         "Limpieza de sillones, colchones y tapizados con maquina extractor"),
  ("Limpieza de vidrios en altura", "Limpieza exterior de ventanas y frentes de edificios con equipo"),
  ("Limpieza de piletas",           "Mantenimiento y limpieza quimica de piletas de natacion"),
  # JARDINERIA
  ("Jardinero",                     "Mantenimiento y diseno de jardines, cesped y plantas"),
  ("Podador de arboles",            "Poda de arboles grandes, palmeras y arbusto en altura"),
  ("Cortador de pasto",             "Corte y mantenimiento de cesped con maquinas propias"),
  ("Instalador de jardines",        "Diseno y construccion de jardines con riego, iluminacion y plantas"),
  ("Instalador de riego automatico", "Diseno e instalacion de sistemas de riego por goteo y aspersion"),
  # PILETAS
  ("Mantenimiento de piletas",      "Control de quimica, limpieza y revision de equipos de natacion"),
  ("Instalador de piletas",         "Construccion e instalacion de piletas de hormigon, fibra y prefabricadas"),
  ("Tecnico en piscinas",           "Reparacion de filtros, bombas, calderas y equipos de piletas"),
  # MUDANZAS Y FLETES
  ("Flete / Mudanza",               "Transporte de muebles y enseres para mudanzas locales y regionales"),
  ("Flete express",                 "Transporte rapido de paquetes, electrodomesticos y objetos varios"),
  # TECNICO EN ELECTRODOMESTICOS
  ("Tecnico en lavarropas",         "Reparacion de lavarropas de carga frontal y superior de todas las marcas"),
  ("Tecnico en heladera",           "Reparacion de heladeras, freezers y equipos de refrigeracion domestica"),
  ("Tecnico en lavavajillas",       "Reparacion de lavavajillas y lavadoras automaticas"),
  ("Tecnico en horno electrico",    "Reparacion de hornos electricos y microondas de uso domestico"),
  ("Tecnico en aire acondicionado", "Instalacion, carga de gas y reparacion de equipos de aire"),
  ("Tecnico en refrigeracion",      "Reparacion de equipos de refrigeracion industrial y camaras de frio"),
  ("Tecnico en extractores",        "Instalacion y reparacion de campanas extractoras y ventiladores"),
  # SEGURIDAD Y AUTOMATIZACION
  ("Instalador de alarmas",         "Colocacion de sistemas de alarma perimetral y de interior"),
  ("Programador de alarmas",        "Configuracion y programacion de centrales de alarma y sensores"),
  ("Instalador de camaras",         "Instalacion de sistemas CCTV, camaras IP y DVR/NVR"),
  ("Instalador de video portero",   "Colocacion de porteros electricos y videoporteros con pantalla"),
  ("Instalador de intercomunicadores", "Instalacion de sistemas de intercomunicacion para edificios"),
  ("Instalador de persianas electricas", "Colocacion de persianas y cortinas motorizadas con automatizacion"),
  # TECNICO EN TECNOLOGIA
  ("Tecnico en PC",                 "Reparacion de computadoras, formateo, instalacion de sistemas y hardware"),
  ("Tecnico en celulares",          "Reparacion de pantallas, bateria y componentes de smartphones"),
  ("Tecnico en TV",                 "Reparacion de televisores LED, OLED y pantallas planas"),
  ("Instalador de antenas",         "Colocacion de antenas de TV abierta, satelital y de radio"),
  ("Instalador de internet",        "Colocacion de routers, extension de redes WiFi y cableado de red"),
  ("Instalador de redes",           "Instalacion de redes LAN, cableado estructurado y switches"),
  ("Instalador de fibra optica",    "Tendido y conexion de fibra optica para internet y comunicaciones"),
  ("Tecnico en audio",              "Instalacion y reparacion de equipos de audio domiciliario y profesional"),
  ("Instalador de home cinema",     "Instalacion de sistemas de cine en casa, proyectores y pantallas"),
  ("Instalador de parlantes",       "Colocacion de sistemas de sonido empotrado y parlantes de exterior"),
  # CORTINAS Y DECORACION
  ("Cortinero",                     "Colocacion de cortinas, rieles, barras y telas en todo tipo de ventanas"),
  ("Instalador de persianas",       "Colocacion de persianas de PVC, madera, aluminio y screen"),
  ("Instalador de cortinas enrollables", "Colocacion de cortinas roller, black out y screen"),
  ("Instalador de toldos",          "Colocacion de toldos retractiles, fijos y caidas de lona"),
  ("Tapicero",                      "Tapizado y retapizado de sillones, sillas, cabeceras y automoviles"),
  # CONSTRUCCIONES EXTERIORES
  ("Instalador de asadores",        "Construccion e instalacion de asadores de ladrillos y acero inoxidable"),
  ("Instalador de quinchos",        "Construccion de quinchos, patio techado y espacios de entretenimiento"),
  ("Instalador de gazebos",         "Colocacion de gazebos, patios cubiertos y pergolas con membrana"),
  # CUIDADO DE PERSONAS Y MASCOTAS
  ("Cuidado de adultos mayores",    "Acompanamiento y asistencia de personas mayores en domicilio"),
  ("Niñera / Babysitter",           "Cuidado de ninos en domicilio por horas o de forma permanente"),
  ("Cuidado de mascotas",           "Cuidado y alimentacion de mascotas en domicilio del dueno"),
  ("Paseador de perros",            "Paseo diario de perros de todos los tamanios con seguro incluido"),
  ("Veterinaria a domicilio",       "Consultas, vacunas y curaciones veterinarias en el domicilio"),
  ("Peluqueria canina a domicilio", "Bano, corte y arreglo de mascotas en el domicilio"),
  # OTROS SERVICIOS
  ("Fumigacion de plagas",          "Fumigacion contra cucarachas, ratas, hormigas, vinchucas y mosquitos"),
  ("Control de termitas",           "Tratamiento preventivo y correctivo contra termitas en madera y estructura"),
  ("Diseñador de interiores",       "Asesoramiento y proyecto de renovacion de ambientes y espacios"),
  ("Arquitecto",                    "Proyecto, direccion de obra y tramites de permisos de construccion"),
  ("Tecnico en ascensores",         "Mantenimiento y reparacion de ascensores y elevadores domiciliarios"),
  ("Mantenimiento edilicio",        "Servicio integral de mantenimiento preventivo de edificios y consorcios"),
  ("Pintor automotriz",             "Pintura y preparacion de carroceria de autos y motos"),
  ("Electrodomesticos en general",  "Reparacion general de equipos electricos y electrodomesticos del hogar"),
  ("Instalador de asfalto",         "Bacheo y pavimentacion de entradas de garaje y senderos privados"),
  ("Instalador de contrapiso",      "Nivelacion y colocacion de contrapiso de cemento y autonivelante"),
  ("Instalador de microcemento",    "Aplicacion de microcemento en pisos, paredes y muebles"),
  ("Pintor de microcemento",        "Aplicacion y pulido de microcemento decorativo en interiores"),
  ("Colocador de cortinas de bano", "Instalacion de cortinas de bano, mamparas y divisores de vidrio"),
  ("Instalador de mamparas",        "Colocacion de mamparas de vidrio para duchas y banos"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Prestadores"

# ── Columnas de prestadores ────────────────────────────────────────────────────
columns = [
    ("nombre",                    "Nombre completo",               22, True),
    ("telefono",                  "Telefono (sin 0 ni 15)",         20, True),
    ("categorias",                "Categorias (separar con ,)",     30, True),
    ("zona_sugerida",             "Zona / Barrio",                  24, False),
    ("estado",                    "Estado",                         14, True),
    ("pausado_hasta",             "Pausado hasta (fecha)",          20, False),
    ("origen",                    "Origen",                        18, True),
    ("quien_recomienda_nombre",   "Recomienda: nombre",            24, False),
    ("quien_recomienda_tel",      "Recomienda: tel",               22, False),
    ("quien_recomienda_barrio",   "Recomienda: barrio",            22, False),
    ("estrellas_iniciales",       "Estrellas 1-5",                 16, False),
    ("descripcion_detallada",     "Descripcion del prestador",     32, False),
    ("dni_numero",                "DNI (numero)",                  18, False),
]

# ── Estilos ────────────────────────────────────────────────────────────────────
COLOR_HEADER   = "1A1A2E"
COLOR_HEADER_TXT = "3CE6C5"
COLOR_EXAMPLE  = "E8F8F6"
COLOR_ROW_ALT  = "F4FFFE"

header_font  = Font(name="Calibri", bold=True, color=COLOR_HEADER_TXT, size=11)
example_font = Font(name="Calibri", color="888888", italic=True, size=10)
body_font    = Font(name="Calibri", size=11)
thin_border  = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)

# ── Fila 1: OBLIGATORIO / opcional ────────────────────────────────────────────
ws.row_dimensions[1].height = 18
for i, (field, label, width, req) in enumerate(columns, 1):
    c = ws.cell(row=1, column=i)
    c.value = "OBLIGATORIO" if req else "opcional"
    c.font  = Font(name="Calibri", size=9, color="3CE6C5" if req else "999999", bold=req)
    c.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center")

# ── Fila 2: etiquetas ──────────────────────────────────────────────────────────
ws.row_dimensions[2].height = 32
for i, (field, label, width, req) in enumerate(columns, 1):
    c = ws.cell(row=2, column=i)
    c.value = label
    c.font  = header_font
    c.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Fila 3: nombre tecnico ─────────────────────────────────────────────────────
ws.row_dimensions[3].height = 16
for i, (field, label, width, req) in enumerate(columns, 1):
    c = ws.cell(row=3, column=i)
    c.value = f"[{field}]"
    c.font  = Font(name="Courier New", size=9, color="999999")
    c.fill  = PatternFill("solid", fgColor="111111")
    c.alignment = Alignment(horizontal="center")

# ── Filas de ejemplo ───────────────────────────────────────────────────────────
examples = [
    ["Juan Perez",   "3511234567", "Plomero, Gasista matriculado", "Nueva Cordoba", "activo",   "", "recomendacion",   "Maria Garcia", "3519876543", "Nueva Cordoba", 5, "Plomero y gasista con 10 anos de experiencia.", "28456789"],
    ["Ana Gonzalez", "3514567890", "Electricista",                 "Alta Cordoba",  "activo",   "", "autoregistro",    "",             "",           "",              4, "Electricista matriculada, especialidad en tableros.", ""],
    ["Carlos Suarez","3516543210", "Gasista matriculado",          "General Paz",   "propuesto","", "ingreso_manual",  "Luis Torres",  "3512345678", "Guemes",        5, "", ""],
]
for r, ex in enumerate(examples, 4):
    ws.row_dimensions[r].height = 22
    for i, v in enumerate(ex, 1):
        c = ws.cell(row=r, column=i)
        c.value = v if v != "" else None
        c.font  = example_font
        c.fill  = PatternFill("solid", fgColor=COLOR_EXAMPLE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border

# ── Filas vacías (7-156) ───────────────────────────────────────────────────────
for row in range(7, 157):
    ws.row_dimensions[row].height = 22
    for col in range(1, len(columns) + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = PatternFill("solid", fgColor="FFFFFF" if row % 2 == 0 else COLOR_ROW_ALT)
        c.font   = body_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border

# ── Validaciones ───────────────────────────────────────────────────────────────
ESTADOS  = "propuesto,activo,inactivo,pausado"
ORIGENES = "recomendacion,autoregistro,ingreso_manual"
ESTRELLAS = "1,2,3,4,5"

def make_dv(formula, col, prompt_title, prompt_body):
    dv = DataValidation(
        type="list", formula1=f'"{formula}"',
        showInputMessage=True, promptTitle=prompt_title, prompt=prompt_body,
        showErrorMessage=True, errorTitle="Invalido", error=f"Opciones: {formula}"
    )
    dv.sqref = f"{col}4:{col}156"
    return dv

# col E = estado, G = origen, K = estrellas  (C = categorias: texto libre, sin dropdown)
ws.add_data_validation(make_dv(ESTADOS,  "E", "Estado",    "Estado del prestador en SERVY"))
ws.add_data_validation(make_dv(ORIGENES, "G", "Origen",    "Como llego este prestador a SERVY"))
ws.add_data_validation(make_dv(ESTRELLAS,"K", "Estrellas", "Puntaje inicial 1 a 5"))

# Fecha pausado_hasta (col F)
dv_fecha = DataValidation(type="date", showInputMessage=True, promptTitle="Fecha", prompt="Formato DD/MM/AAAA")
dv_fecha.sqref = "F4:F156"
ws.add_data_validation(dv_fecha)

# Hint en col C (categorias — texto libre, sin restriccion de lista)
dv_cat_hint = DataValidation(
    type="textLength", operator="greaterThanOrEqual", formula1="0",
    showInputMessage=True,
    promptTitle="Categorias",
    prompt="Una o mas categorias separadas por coma. Ej: Plomero, Gasista matriculado"
)
dv_cat_hint.sqref = "C4:C156"
ws.add_data_validation(dv_cat_hint)

ws.freeze_panes = "A4"

# ── Hoja: LISTAS DE REFERENCIA ─────────────────────────────────────────────────
ws_ref = wb.create_sheet("Listas de referencia")
ws_ref.column_dimensions["A"].width = 34
ws_ref.column_dimensions["B"].width = 60
ws_ref.column_dimensions["C"].width = 18
ws_ref.column_dimensions["D"].width = 22

def ref_header(ws, col, title):
    c = ws[f"{col}1"]
    c.value = title
    c.font  = Font(bold=True, color="FFFFFF", size=11)
    c.fill  = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center")

ref_header(ws_ref, "A", "CATEGORIAS / OFICIOS")
ref_header(ws_ref, "B", "DESCRIPCION (para tabla categorias en Supabase)")
ref_header(ws_ref, "C", "ESTADOS")
ref_header(ws_ref, "D", "ORIGENES")

ws_ref.row_dimensions[1].height = 24

for i, (nombre, desc) in enumerate(CATEGORIAS, 2):
    ws_ref.row_dimensions[i].height = 20
    cn = ws_ref.cell(row=i, column=1, value=nombre)
    cd = ws_ref.cell(row=i, column=2, value=desc)
    cn.font = Font(name="Calibri", size=10)
    cd.font = Font(name="Calibri", size=10, color="555555")
    color = "FFFFFF" if i % 2 == 0 else "F4FFFE"
    cn.fill = cd.fill = PatternFill("solid", fgColor=color)

for i, v in enumerate(ESTADOS.split(","), 2):
    ws_ref.cell(row=i, column=3, value=v.strip())
for i, v in enumerate(ORIGENES.split(","), 2):
    ws_ref.cell(row=i, column=4, value=v.strip())

# ── Guardar ────────────────────────────────────────────────────────────────────
output = "prestadores_servy_v2.xlsx"
wb.save(output)
print("Archivo generado: " + output)
print("  -> Elimina las filas 4-6 (ejemplos) antes de importar a Supabase.")
print("  -> La columna Categorias acepta multiples valores separados por coma.")
print("  -> Total de categorias en hoja de referencia: " + str(len(CATEGORIAS)))
